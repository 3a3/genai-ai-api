"""Excel (xlsx) ローダー。

QA データを想定:
- 1 行 = 1 Section
- 期待カラム: "question" / "answer"（または "Q" / "A"）
- 見つからなければ全カラムを連結して text として扱う
"""

from __future__ import annotations

import io

from ..errors import CorruptedFileError, NoTextExtractedError, PasswordProtectedError
from ..types import Section

_QUESTION_KEYS = {"question", "q", "質問", "問い"}
_ANSWER_KEYS = {"answer", "a", "回答", "答え"}


def _norm(s: str) -> str:
    return s.strip().lower()


def load_xlsx(blob_path: str, content: bytes) -> list[Section]:
    try:
        import openpyxl  # 遅延 import
    except ImportError as exc:  # pragma: no cover
        raise CorruptedFileError(blob_path, f"openpyxl is not installed: {exc}") from exc

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg or "bad zip" in msg:
            raise PasswordProtectedError(blob_path, f"xlsx may be password protected: {exc}") from exc
        raise CorruptedFileError(blob_path, f"failed to open xlsx: {exc}") from exc

    sections: list[Section] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        headers = [str(c).strip() if c is not None else "" for c in header_row]
        norm_headers = [_norm(h) for h in headers]

        q_idx = next((i for i, h in enumerate(norm_headers) if h in _QUESTION_KEYS), None)
        a_idx = next((i for i, h in enumerate(norm_headers) if h in _ANSWER_KEYS), None)

        for row_idx, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            values = ["" if v is None else str(v).strip() for v in row]
            if not any(values):
                continue

            if q_idx is not None and a_idx is not None and q_idx < len(values) and a_idx < len(values):
                question = values[q_idx]
                answer = values[a_idx]
                if not (question or answer):
                    continue
                text = f"Q: {question}\nA: {answer}"
                title = question[:80] if question else f"{sheet_name} row {row_idx}"
            else:
                # ヘッダ無し / 列が見つからない場合: ヘッダ名: 値 を連結
                parts = [f"{headers[i] or f'col{i+1}'}: {values[i]}" for i in range(len(values)) if values[i]]
                if not parts:
                    continue
                text = "\n".join(parts)
                title = values[0][:80] if values[0] else f"{sheet_name} row {row_idx}"

            sections.append(
                Section(
                    text=text,
                    title=title,
                    section="",
                    page=None,
                    locator=f"sheet={sheet_name}#row={row_idx}",
                )
            )

    wb.close()

    if not sections:
        raise NoTextExtractedError(blob_path, "no rows extracted from xlsx")

    return sections
